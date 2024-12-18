import openpyxl
from openpyxl.workbook import Workbook



class ExcelFunction:
    def __init__(self, driver):
        self.driver = driver

    @staticmethod
    def get_row_count(path,sheet_name):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def get_column_count(path,sheet_name):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        return sheet.max_column

    @staticmethod
    def read_data(path,sheet_name,row_num,colum_num):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        return sheet.cell(row = row_num, column = colum_num).value

    @staticmethod
    def write_data(path,sheet_name,row_num,colum_num,data):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        sheet.cell(row = row_num, column = colum_num).value = data
        workbook.save(path)

    